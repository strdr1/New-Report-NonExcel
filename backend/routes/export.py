from io import BytesIO
from datetime import datetime
from flask import Blueprint, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from backend.database import get_db
from backend.models import Profile, Year, Month, DailyRecord
from backend.services.calculation_service import CalculationService
from backend.routes.months import _get_initial_values, _fuel_rate

bp = Blueprint('export', __name__, url_prefix='/api/profiles')

MONTH_NAMES_RU = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                  'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
DAYS_RU = ['Вс', 'Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб']

THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_FILL = PatternFill('solid', fgColor='E3F2FD')
TOTAL_FILL = PatternFill('solid', fgColor='FFF9C4')


def _write_header(ws, profile, year, month_num, title, start_row=1):
    mn = MONTH_NAMES_RU[month_num].lower()
    ws.cell(row=start_row, column=1, value=title).font = Font(size=16, bold=True)
    ws.cell(row=start_row + 1, column=1,
            value=f'за {mn} месяц {year} года').font = Font(size=11, bold=True)
    ws.cell(row=start_row + 2, column=1,
            value=f'Автомашина: {profile.car_brand}   Гос. номер: {profile.license_plate}   Водитель: {profile.name}'
            ).font = Font(size=11)
    return start_row + 4


def _autofit(ws, min_width=8, max_width=22):
    for col in ws.columns:
        try:
            letter = get_column_letter(col[0].column)
        except Exception:
            continue
        length = min_width
        for c in col:
            v = c.value
            if v is None:
                continue
            length = max(length, len(str(v)) + 2)
        ws.column_dimensions[letter].width = min(length, max_width)


@bp.route('/<int:profile_id>/years/<int:year_value>/months/<int:month_num>/export', methods=['GET'])
def export_month(profile_id, year_value, month_num):
    db = next(get_db())
    try:
        profile = db.query(Profile).filter(Profile.id == profile_id).first()
        if not profile:
            return jsonify({'error': 'Not Found', 'message': 'Профиль не найден'}), 404

        year = db.query(Year).filter(Year.profile_id == profile_id, Year.year == year_value).first()
        if not year:
            return jsonify({'error': 'Not Found', 'message': f'Год {year_value} не найден'}), 404

        month = db.query(Month).filter(Month.year_id == year.id, Month.month == month_num).first()
        if not month:
            return jsonify({'error': 'Not Found', 'message': f'Месяц {month_num} не найден'}), 404

        initial_odometer, initial_fuel = _get_initial_values(db, year, month_num)
        fuel_rate = _fuel_rate(profile, month_num, month)

        records = db.query(DailyRecord).filter(
            DailyRecord.month_id == month.id
        ).order_by(DailyRecord.day).all()

        wb = Workbook()

        # ---------- Лист 1: Путевые записи ----------
        ws1 = wb.active
        ws1.title = 'Путевые записи'
        row = _write_header(ws1, profile, year, month_num, 'ПУТЕВЫЕ ЗАПИСИ')

        headers = ['Дата', 'День нед.', 'Спидометр конец дня', 'Пробег путевка',
                   'Бензин путевка', 'Спидометр начало', 'Спидометр конец',
                   'Бензин остаток', 'Бензин получено']
        for i, h in enumerate(headers, 1):
            c = ws1.cell(row=row, column=i, value=h)
            c.font = Font(bold=True)
            c.alignment = CENTER
            c.border = BORDER
            c.fill = HEADER_FILL
        ws1.row_dimensions[row].height = 32
        row += 1

        for rec in records:
            d = datetime(year_value, month_num, rec.day)
            dow = DAYS_RU[(d.weekday() + 1) % 7]
            vals = [
                d.strftime('%d.%m.%Y'),
                dow,
                rec.odometer_end_day,
                rec.distance_km,
                rec.fuel_waybill,
                rec.odometer_start,
                rec.odometer_end,
                rec.fuel_remaining,
                rec.fuel_received,
            ]
            for i, v in enumerate(vals, 1):
                c = ws1.cell(row=row, column=i, value=v)
                c.alignment = CENTER
                c.border = BORDER
            row += 1

        _autofit(ws1)

        # ---------- Лист 2: Отчёт ----------
        ws2 = wb.create_sheet('Отчёт')
        row = _write_header(ws2, profile, year, month_num, 'ОТЧЕТ')

        # Двойной заголовок
        ws2.cell(row=row, column=1, value='Дата')
        ws2.cell(row=row, column=2, value='Начало дня')
        ws2.cell(row=row, column=4, value='Конец дня')
        ws2.cell(row=row, column=6, value='Получено бензина')
        ws2.cell(row=row, column=7, value='За день')
        ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws2.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
        ws2.merge_cells(start_row=row, start_column=6, end_row=row + 1, end_column=6)

        sub = ['', 'км', 'л', 'км', 'л', '', 'км', 'л']
        for i, s in enumerate(sub, 1):
            ws2.cell(row=row + 1, column=i, value=s)

        for r in (row, row + 1):
            for col in range(1, 9):
                c = ws2.cell(row=r, column=col)
                c.font = Font(bold=True)
                c.alignment = CENTER
                c.border = BORDER
                c.fill = HEADER_FILL
        ws2.row_dimensions[row].height = 22
        row += 2

        max_odometer_end = initial_odometer
        prev_fuel_remaining = initial_fuel
        tot_rec = tot_km = tot_fuel = 0.0

        for rec in records:
            km_za_den = CalculationService.calculate_km_za_den(
                rec.odometer_end_day, rec.distance_km, max_odometer_end
            )
            fuel_waybill = CalculationService.calculate_fuel_waybill(km_za_den, fuel_rate)
            fuel_remaining = CalculationService.calculate_fuel_remaining(
                prev_fuel_remaining, fuel_waybill, rec.fuel_received, km_za_den
            )

            odometer_end_report = max_odometer_end + km_za_den if km_za_den > 0 else None
            d = datetime(year_value, month_num, rec.day)
            vals = [
                d.strftime('%d.%m.%Y'),
                round(max_odometer_end, 0),
                round(prev_fuel_remaining, 2) if prev_fuel_remaining else None,
                round(odometer_end_report, 0) if odometer_end_report else None,
                round(fuel_remaining, 2) if fuel_remaining else None,
                rec.fuel_received,
                round(km_za_den, 0) if km_za_den > 0 else None,
                round(fuel_waybill, 3) if km_za_den > 0 else None,
            ]
            for i, v in enumerate(vals, 1):
                c = ws2.cell(row=row, column=i, value=v)
                c.alignment = CENTER
                c.border = BORDER
            row += 1

            if rec.fuel_received:
                tot_rec += rec.fuel_received
            if km_za_den > 0:
                tot_km += km_za_den
                tot_fuel += fuel_waybill

            if odometer_end_report and odometer_end_report > max_odometer_end:
                max_odometer_end = odometer_end_report
            if km_za_den > 0:
                prev_fuel_remaining = fuel_remaining

        # ИТОГО
        ws2.cell(row=row, column=1, value='ИТОГО')
        ws2.cell(row=row, column=6, value=round(tot_rec, 0) if tot_rec else None)
        ws2.cell(row=row, column=7, value=round(tot_km, 0) if tot_km else None)
        ws2.cell(row=row, column=8, value=round(tot_fuel, 3) if tot_fuel else None)
        for col in range(1, 9):
            c = ws2.cell(row=row, column=col)
            c.font = Font(bold=True)
            c.alignment = CENTER
            c.border = BORDER
            c.fill = TOTAL_FILL

        _autofit(ws2)

        # Сохраняем в память
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'{profile.name}_{MONTH_NAMES_RU[month_num]}_{year_value}.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    finally:
        db.close()
